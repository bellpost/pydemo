#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2017/6/27 22:19
# @Author  : Bellpost
# @File    : deal_excel.py
#
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Color
from openpyxl.styles import colors,Alignment,PatternFill
import datetime
import time,os

class Deal_excel:
    def __init__(self, *args, **kwargs):
        self._fields = ['name', 'title', 'index']
        self.wb = Workbook()
        if len(args) != len(self._fields):
            raise TypeError('Expected {} arguments'.format(len(self._fields)))
        # set the arguments
        for name, value in zip(self._fields, args):
            setattr(self, name, value)

        extra_args = kwargs.keys() - self._fields
        for name in extra_args:
            setattr(self, name, kwargs.pop(name))
        if kwargs:
            raise TypeError('Duplicate value for {} '.format(','.join(kwargs)))
        self.dataPjo = self.wb.create_sheet(self.title, int(self.index))

    def init_excel(self,title,index):
        self.dataPjo2=self.wb.create_sheet(title, int(index))

    def paser_count_rows(self,rows):
        for i in rows:
            self.dataPjo2.append(i)

    def paser_count(self,headers,rows):
        self.dataPjo.append(headers)
        for i in rows:
            self.dataPjo.append(i)

    def save_excel(self):
        self.wb.save(self.name+'.xlsx')


if __name__ == '__main__':
    headers = ['Symbol', 'Price', 'Date', 'Time', 'Change', 'Volume']
    rows = [('AA', 39.48, '6/11/2007', '9:36am', -0.18, 181800),
            ('AIG', 71.38, '6/11/2007', '9:36am', -0.15, 195500),
            ('AXP', 62.58, '6/11/2007', '9:36am', -0.46, 935000),
            ]
    # a = Deal_excel('数据报表_'+time.strftime('%Y-%m-%d'),'DateProject','0',A = 'c')
    # a.paser_count(headers,rows)
    # a.init_excel('csv',1)
    # import csv
    # with open('/data/GPS_DATA/path.csv','r') as f:
    #     f_csv=csv.reader(f)
    #     print(f_csv)
    #     a.paser_count_rows(f_csv)
    # # os.chdir('/data/GPS_DATA/')
    # a.save_excel()

    # appSer= Deal_excel('巡检报表_'+time.strftime('%Y-%m-%d'),'ApplicationServer','0')
    # ws = appSer.dataPjo
    # ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=7)
    # # titlecell = ws['A1']
    # ft = Font(name=u'宋体',size=16,color=colors.BLACK,bold=True)
    # ft1 = Font(name=u'宋体', bold=True)
    # ftwarn = Font(name=u'宋体', color=colors.RED, bold=True)
    # fill = PatternFill(fill_type='solid',start_color = '00BFFF', end_color = '00BFFF')
    # align=Alignment(horizontal='center', vertical='center')
    # titlecell=ws.cell(row=1, column=1, value='系统巡检报表')
    # titlecell.font = ft
    # titlecell.alignment=align
    # ws.cell(row=2, column=1, value='检查人：').font=ft1
    # ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=7)
    # checkman =ws.cell(row=2, column=2, value='钟博文')
    # ws.cell(row=3, column=1, value='检查时间：').font=ft1
    # ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=7)
    # checktime = ws.cell(row=3, column=2, value=datetime.datetime.now()).alignment=align
    # ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=7)
    # ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=7)
    # systeminfo = ws.cell(row=4, column=1, value='系统信息')
    # systeminfo.font = ft1
    # systeminfo.fill = fill
    # systeminfo.alignment = align
    # ws.cell(row=5, column=1, value='检查IP').font=ft1
    # ws.cell(row=5, column=2, value='检查点').font=ft1
    # ws.cell(row=5, column=3, value='检查细项').font=ft1
    # ws.cell(row=5, column=4, value='检查情况').font=ft1
    # ws.cell(row=5, column=5, value='基准').font=ft1
    # ws.cell(row=5, column=6, value='是否异常').font=ft1
    # ws.cell(row=5, column=7, value='备注').font=ft1
    # appSer.save_excel()'

    wb = load_workbook("巡检报表_2017-07-14.xlsx")
    ws = wb.get_sheet_by_name('ApplicationServer')

    ws_row_len = ws.max_row #ws.rows
    ws_columns_len = ws.max_column
    #print(ws_row_len,ws_columns_len)
    align = Alignment(horizontal='center', vertical='center')
    start_row=6
    stop_row=ws_row_len
    start_column = 'a'
    stop_column = 'b'
    for cloumn_num in range(ord(start_column),ord(stop_column)+1):
        n = 0
        for i in range(start_row,stop_row+1):
            # n=i
            # while ws[chr(cloumn_num)+str(n)].value is None:
            #     n=i-1
            if ws[chr(cloumn_num)+str(i)].value == ws[chr(cloumn_num)+str(i+1)].value :
                n=n+1
                print(n)
            else:
                start_temp_row = i - n
                stop_temp_row = i
                print('start_temp_row:' + str(start_temp_row), 'stop_temp_row:' + str(stop_temp_row),
                      'cloumn_num:' + str(cloumn_num))
                ws.merge_cells(start_row=start_temp_row, start_column=cloumn_num-96, end_row=stop_temp_row, end_column=cloumn_num-96)
                ws.cell(row=start_temp_row, column=cloumn_num-96).alignment=align
                n=0

    wb.save('temp.xlsx')










