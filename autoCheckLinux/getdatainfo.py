#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2017/7/13 11:16
# @Author  : Bellpost
# @File    : getdatainfo.py
#

from fabric.api import *
from datetime import datetime
import time,datetime
import logging
import csv

from openpyxl import load_workbook
from openpyxl.styles import Font, Color
from openpyxl.styles import colors,Alignment,PatternFill
import datetime
import time,os
from config import dealconfig
from deal_excel import Deal_excel

logger = logging.getLogger()
fh = logging.FileHandler('getdatainfo.log')
formatter = logging.Formatter('%(message)s')
fh.setFormatter(formatter)
logger.addHandler(fh)
logger.setLevel(logging.ERROR)

config_file_path="./conf/config.ini"
dealconfig = dealconfig(config_file_path)
diskspaceroot=dealconfig.get('systembaseinfo','root')
diskspace_sda1=dealconfig.get('systembaseinfo','diskspace_sda1')
diskspace_sdb1=dealconfig.get('systembaseinfo','diskspace_sdb1')
diskspace_home=dealconfig.get('systembaseinfo','diskspace_home')
memery = dealconfig.get('systembaseinfo','memery')
swap = dealconfig.get('systembaseinfo','swap')

serverdict={
'192.168.130.XX':'暂无应用',
'192.168.130.XX':'暂无应用'
}


env.user = 'user'
env.roledefs = {
	'tocc': [''],
	'jks': [''],
	'zs' : [''],
    'jks_245': ['']
    }
env.passwords = {
    'user@IP:22':'passwd',
    }
env.hosts = ['']
contentall = []


def getdata():
    rows = local("cat /home/tocc/dataToCsv/zspath.csv",capture=True)
    local("echo '' > /home/tocc/dataToCsv/zspath.csv")
    return rows
def getcont():
    rows = local("cat /opt/xjpy/zs_baobiao/zhishu.csv",capture=True)
    return rows
def getcont1():
    rows = local("cat /opt/xjpy/zs_baobiao/zhishudata.csv",capture=True)
    return rows
def getdiskdata():
    rows = local("cat /home/dataToCsv/zsdiskdataPath.csv",capture=True)
    local("echo '' > /home/dataToCsv/zsdiskdataPath.csv")
    return rows

@task
@hosts('172.28.203.135')
def getpathdata():
    rows = run("cat /data/GPS_DATA/path.csv")
    run("echo '' > /data/GPS_DATA/path.csv")
    return rows




def create_excel():
    appSer= Deal_excel('zs巡检报表_temp_'+time.strftime('%Y-%m-%d'),'服务器巡检','0')
    ws = appSer.dataPjo
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=8)
    # titlecell = ws['A1']
    ft = Font(name=u'宋体',size=16,color=colors.BLACK,bold=True)
    ft1 = Font(name=u'宋体', bold=True)
    ftwarn = Font(name=u'宋体', color=colors.RED, bold=True)
    fill = PatternFill(fill_type='solid',start_color = '00BFFF', end_color = '00BFFF')
    align=Alignment(horizontal='center', vertical='center')
    titlecell=ws.cell(row=1, column=1, value='表1 服务器使用情况巡检表（日报）')
    titlecell.font = ft
    titlecell.alignment=align
    ws.cell(row=2, column=1, value='检查人：').font=ft1
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=8)
    checkman =ws.cell(row=2, column=2, value='钟博文')
    ws.cell(row=3, column=1, value='检查时间：').font=ft1
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=8)
    checktime = ws.cell(row=3, column=2, value=datetime.datetime.now()).alignment=align
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=8)
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=8)
    systeminfo = ws.cell(row=4, column=1, value='服务器巡检结果')
    systeminfo.font = ft1
    systeminfo.fill = fill
    systeminfo.alignment = align
    ws.cell(row=5, column=1, value='服务器IP').font=ft1
    ws.cell(row=5, column=2, value='用途').font=ft1
    ws.cell(row=5, column=3, value='检查点').font=ft1
    ws.cell(row=5, column=4, value='检查项').font=ft1
    ws.cell(row=5, column=5, value='检查情况').font=ft1
    ws.cell(row=5, column=6, value='基准').font=ft1
    ws.cell(row=5, column=7, value='是否正常').font=ft1
    ws.cell(row=5, column=8, value='备注').font=ft1
    for i in contentall:
        ws.append(i)
    contentall.clear()
    diskfile = getdiskdata()
    for cell in diskfile.split('\n'):
        ws.append(cell.split(','))
    appSer.init_excel('数据巡检',1)
    ws1 = appSer.dataPjo2
    ws1.merge_cells(start_row=1,start_column=1,end_row=1,end_column=7)
    titlecell1=ws1.cell(row=1, column=1, value='数据巡检报表')
    titlecell1.font = ft
    titlecell1.alignment=align
    ws1.cell(row=2, column=1, value='检查人：').font=ft1
    ws1.merge_cells(start_row=2, start_column=2, end_row=2, end_column=7)
    checkman1 =ws1.cell(row=2, column=2, value='钟博文')
    ws1.cell(row=3, column=1, value='检查时间：').font=ft1
    ws1.merge_cells(start_row=3, start_column=2, end_row=3, end_column=7)
    checktime1 = ws1.cell(row=3, column=2, value=datetime.datetime.now()).alignment=align
    ws1.merge_cells(start_row=3, start_column=2, end_row=3, end_column=7)
    ws1.merge_cells(start_row=4, start_column=1, end_row=4, end_column=7)
    dataminfo = ws1.cell(row=4, column=1, value='数据信息')
    dataminfo.font = ft1
    dataminfo.fill = fill
    dataminfo.alignment = align
    ws1.cell(row=5, column=1, value='检查IP').font=ft1
    ws1.cell(row=5, column=2, value='服务器作用').font=ft1
    ws1.cell(row=5, column=3, value='数据项').font = ft1
    ws1.cell(row=5, column=4, value='时间').font=ft1
    ws1.cell(row=5, column=5, value='检查情况').font=ft1
    ws1.cell(row=5, column=6, value='是否正常').font=ft1
    ws1.cell(row=5, column=7, value='备注').font=ft1
    contfile=getdata()
   # for cont in contfile.values():
    for cell in  contfile.split('\n'):
        ws1.append(cell.split(','))
    lostfile = execute(getpathdata)
    for cont1 in lostfile.values():
        for cell1 in  cont1.split('\n'):
            ws1.append(cell1.split(','))
    appSer.init_excel('系统巡检',2)
    ws2 = appSer.dataPjo2
    ws2.merge_cells(start_row=1,start_column=1,end_row=1,end_column=7)
    titlecell1=ws2.cell(row=1, column=1, value='表3 软件系统巡检表（日报）')
    titlecell1.font = ft
    titlecell1.alignment=align
    ws2.cell(row=2, column=1, value='检查人：').font=ft1
    ws2.merge_cells(start_row=2, start_column=2, end_row=2, end_column=7)
    checkman1 =ws2.cell(row=2, column=2, value='钟博文')
    ws2.cell(row=3, column=1, value='检查时间：').font=ft1
    ws2.merge_cells(start_row=3, start_column=2, end_row=3, end_column=7)
    checktime1 = ws2.cell(row=3, column=2, value=datetime.datetime.now()).alignment=align
    ws2.merge_cells(start_row=3, start_column=2, end_row=3, end_column=7)
    ws2.merge_cells(start_row=4, start_column=1, end_row=4, end_column=7)
    dataminfo = ws2.cell(row=4, column=1, value='系统信息')
    dataminfo.font = ft1
    dataminfo.fill = fill
    dataminfo.alignment = align
    ws2.cell(row=5, column=1, value='系统名称').font=ft1
    ws2.cell(row=5, column=2, value='检查项').font=ft1
    ws2.cell(row=5, column=3, value='显示正常').font=ft1
    ws2.cell(row=5, column=4, value='数据正常').font=ft1
    ws2.cell(row=5, column=5, value='解决情况').font=ft1
    ws2.cell(row=5, column=6, value='备注').font=ft1
    contfile2=getcont()
    #for cont2 in contfile2.values():
    for cell2 in  contfile2.split('\n'):
        ws2.append(cell2.split(','))
    appSer.save_excel()
    appSer.init_excel('数据巡检1',3)
    ws3 = appSer.dataPjo2
    ws3.merge_cells(start_row=1,start_column=1,end_row=1,end_column=7)
    titlecell1=ws3.cell(row=1, column=1, value='表2 数据中心数据巡检表（日报）')
    titlecell1.font = ft
    titlecell1.alignment=align
    ws3.cell(row=2, column=1, value='检查人：').font=ft1
    ws3.merge_cells(start_row=2, start_column=2, end_row=2, end_column=7)
    checkman1 =ws3.cell(row=2, column=2, value='钟博文')
    ws3.cell(row=3, column=1, value='检查时间：').font=ft1
    ws3.merge_cells(start_row=3, start_column=2, end_row=3, end_column=7)
    checktime1 = ws3.cell(row=3, column=2, value=datetime.datetime.now()).alignment=align
    ws3.merge_cells(start_row=3, start_column=2, end_row=3, end_column=7)
    ws3.merge_cells(start_row=4, start_column=1, end_row=4, end_column=7)
    dataminfo = ws3.cell(row=4, column=1, value='数据巡检结果')
    dataminfo.font = ft1
    dataminfo.fill = fill
    dataminfo.alignment = align
    ws3.cell(row=5, column=1, value='类型').font=ft1
    ws3.cell(row=5, column=2, value='数据项').font=ft1
    ws3.cell(row=5, column=3, value='时间').font=ft1
    ws3.cell(row=5, column=4, value='是否正常').font=ft1
    ws3.cell(row=5, column=5, value='检查情况').font=ft1
    ws3.cell(row=5, column=6, value='备注').font=ft1
    contfile2=getcont1()
    #for cont2 in contfile2.values():
    for cell2 in  contfile2.split('\n'):
        ws3.append(cell2.split(','))
    appSer.save_excel()

@task
@hosts(env.hosts)
#@hosts('172.28.203.135')
def getInfo():
    infofiles=['./conf/systembaseinfo.cmd']
    #判断版本lsb_release -a|grep 'Release'| awk '{print $2}' |cut -d '.' -f 1
    #release=" more /proc/version | awk '{print $3}' |cut -d '.' -f 4 "
    release = " more /proc/version | awk '{print $3}' |grep 'el7' |wc -l "
    if run(release) == '1':
        get_ip = "ifconfig|grep 'inet'|grep -v '127\.0\.0\.1'|grep -v '172\.17\.0\.1'|cut -d: -f2|awk '{print $2}'"
    else:
    #get_ip = "ip -f inet addr | grep -v 127.0.0.1 |  grep inet | awk '{print $NF,$2}' | tr '\n' ',' | sed 's/,$//'"
        get_ip = "ifconfig  | grep 'inet addr:'| grep -v '127.0.0.1' |cut -d: -f2 | awk '{ print $1}'"
    #check_ip='172.28.203.144'
    check_ip = run(get_ip)
    for parent, dirnames, filenames in os.walk('./conf/hosts/'):
        if check_ip in filenames:
            infofiles.append('./conf/hosts/'+check_ip)
    for file in infofiles:
        command_file = open(file, 'r', encoding='gbk')
        for line in command_file.readlines():
            content = []
            line = line.strip('\n')
            items = line.split('#')
            check_type = items[0].encode('gbk')
            check_xx = items[1].encode('gbk')
            check_cmd = items[2]
            check_base_line = items[3]
            opration = check_base_line[0:1]
            base_value = check_base_line[1:]
            result = str(run(check_cmd))
            error_flag = u'正常'
            if result is None or result == '':
                error_flag=u'该项无结果'
                logging.error(datetime.datetime.now().strftime("%Y%m%d")+":"+error_flag+":"+check_ip+":"+check_cmd)
                continue
            else:
                content.append(check_ip)
                content.append(serverdict[check_ip])
                content.append(check_type)
                content.append(check_xx)
                content.append(result)
                content.append(check_base_line)
                if isinstance(result,str) :
                    if "%" in result:
                        result=int(result.replace('%','').strip())
                    else:
                        result = int(result)
                if isinstance(base_value, str):
                    if "%" in base_value:
                        base_value=int(base_value.replace('%',''))
                    else:
                        base_value = int(base_value)
                if opration == '>':
                    if base_value > result:
                        error_flag = u'异常'
                elif opration == '=':
                    if result != base_value:
                        error_flag = u'异常'
                elif opration == '<':
                    if base_value < result:
                        error_flag = u'异常'

                content.append(error_flag)
            contentall.append(tuple(content))

def excel_format(old_excel,sheetname,stop_column):
    wb = load_workbook(old_excel+'.xlsx')
    ws = wb.get_sheet_by_name(sheetname)

    ws_row_len = ws.max_row
    ws_columns_len = ws.max_column
    align = Alignment(horizontal='center', vertical='center')
    start_row=6
    stop_row=ws_row_len
    start_column = 'a'
    if not stop_column :
        stop_column = 'c'
    for cloumn_num in range(ord(start_column),ord(stop_column)+1):
        n = 0
        for i in range(start_row,stop_row+1):
            if ws[chr(cloumn_num)+str(i)].value == ws[chr(cloumn_num)+str(i+1)].value :
                n=n+1
                #print(n)
            else:
                start_temp_row = i - n
                stop_temp_row = i
                if n!=0:
                #print('start_temp_row:' + str(start_temp_row), 'stop_temp_row:' + str(stop_temp_row),'cloumn_num:' + str(cloumn_num))
                    ws.merge_cells(start_row=start_temp_row, start_column=cloumn_num-96, end_row=stop_temp_row, end_column=cloumn_num-96)
                ws.cell(row=start_temp_row, column=cloumn_num-96).alignment=align
                n=0

    wb.save('巡检报表_'+time.strftime('%Y-%m-%d')+'.xlsx')




if __name__ =='__main__':
    execute(getInfo)
    create_excel()
    excel_format('巡检报表_temp_'+time.strftime('%Y-%m-%d'),'服务器巡检','c')
    excel_format('巡检报表_'+time.strftime('%Y-%m-%d'),'数据巡检','b')
    excel_format('巡检报表_' + time.strftime('%Y-%m-%d'), '系统巡检','b')




